# library import
from sklearn.metrics import classification_report, confusion_matrix, accuracy_score
from sklearn.feature_extraction.text import CountVectorizer
import numpy as np
import re
import nltk
from openpyxl import load_workbook
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords  # bibliotēka stop vārdiem/pieturvārdiem
from sklearn import preprocessing
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from nltk.corpus import stopwords
import pickle
from typing import Dict, List, Tuple
import collections

print (stopwords.words('english'))
def text2bow(words: List[str], dictionary: Dict[str, int]) -> List[Tuple[int, int]]:
    word_frequences = collections.defaultdict(int)
    for word in words:
        if word not in dictionary:
            dictionary[word] = len(dictionary)
        word_frequences[dictionary[word]] += 1
    return list(word_frequences.items())


def tokenize_sentences(sentences):
    words = []
    for sentence in sentences:
        w = extract_words(sentence)
        words.extend(w)

    words = sorted(list(set(words)))
    
    
    return words

def extract_words(sentence):
    ignore_words = ['a']
    # nltk.word_tokenize(sentence)
    words = re.sub("[^w]", " ",  sentence).split()
    words_cleaned = [w.lower() for w in words if w not in ignore_words]
    
    return words_cleaned


def bagofwords(sentence, words):
    sentence_words = extract_words(sentence)
    # frequency word count
    bag = np.zeros(len(words))
    for sw in sentence_words:
        for i, word in enumerate(words):
            if word == sw:
                bag[i] += 1
    
    return np.array(bag)


# workbook in the same folder just point out file name
workbook = load_workbook('data.xlsx')
# sheet name programm will take data
worksheet = workbook.get_sheet_by_name('data')
# define
c = 1
# stop words define
stop_words = set(stopwords.words('english'))
# import the necessary module
# create the Labelencoder object
le = preprocessing.LabelEncoder()
# convert the categorical columns into numeric


categories = []
data = []
dictionary = {}
data_arr = []
data_arr2 = []
# for loop, each row separately
for i in worksheet:
    text = worksheet.cell(row=c, column=2).value
    category = worksheet.cell(row=c, column=1).value
    categories.__iadd__([category])
    print(text)
    # remove stop words
    text_without_stop_words = [
        word for word in text.split() if word.lower() not in stop_words]
    print(text_without_stop_words)
    new_text = " ".join(text_without_stop_words)
    print(new_text)
    dataset = nltk.sent_tokenize(new_text, language="english")
    print(dataset)
    for j in range(len(dataset)):
        dataset[j] = dataset[j].lower()  # lowercase
        dataset[j] = re.sub(r'\W', ' ', dataset[j])
        dataset[j] = re.sub(r'\s+', ' ', dataset[j])
        dataset[j] = ''.join([i for i in dataset[j] if not i.isdigit()])

    if(text is None):
        print("null")
    else:
        # print(dataset)
        # pievienojam katru rindu kā nākamo masīva elementu
        data_arr2.__iadd__(dataset)
    # katru jaunu vārdu ar savu id jāpievieno kopīgjaam sarakstam
    data_arr.__iadd__([text2bow(new_text.split(), dictionary)])

    c = c+1
x = len(data_arr2)
print(x)


# categories to number and array
encoded_value = le.fit_transform(categories)  # pārvēršam kategorijas skaitļos
print(encoded_value)
sentences = data_arr2
vocabulary = tokenize_sentences(sentences)
# print(dictionary)


vectorizer = CountVectorizer(analyzer="word", tokenizer=None,
                             preprocessor=None, stop_words=None, max_features=5000)
train_data_features = vectorizer.fit_transform(sentences)
vectoried = vectorizer.transform(sentences).toarray()
print(vectorizer.transform(sentences).toarray())


# train and test data
X, y = vectoried, encoded_value
X_train, X_test, y_train, y_test = train_test_split(
    X, y, test_size=0.2, random_state=1)
# print(X_train)
# print(y_train)
# print(X_test)
print(y_test)
print( X_test, y_test)

x = len(X_train)
y = len(y_train)
x2 = len(X_test)
y2 = len(y_test)
print(x, y)
print(x2, y2)
classifier = RandomForestClassifier(n_estimators=5000, random_state=5)
classifier.fit(X_train, y_train)
y_pred = classifier.predict(X_test)
print(y_train)
print(y_pred)
print(y_test)
# accuracy
print(confusion_matrix(y_test, y_pred))
print(classification_report(y_test, y_pred))
print(accuracy_score(y_test, y_pred))
# model saving in folder
with open('text_classifier', 'wb') as picklefile:
    pickle.dump(classifier, picklefile)
